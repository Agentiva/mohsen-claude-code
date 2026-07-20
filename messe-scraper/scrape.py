#!/usr/bin/env python3
"""
Messe-Aussteller-Scraper
========================

Scrapt die Aussteller (Firmen) der in fairs.json konfigurierten Messen und
schreibt pro Messe eine CSV sowie eine Gesamtdatei (CSV + XLSX).

Strategie ("20-Jahre-Erfahrung"-Ansatz):
  1. method=download  -> direkte Aussteller-Datei (XLS/CSV/PDF) laden & parsen.
     Das ist der zuverlaessigste Weg (Messe-Berlin "Virtual Market Place"
     bietet z.B. eine fertige Ausstellerliste.xls).
  2. method=browser   -> Seite mit Playwright rendern und dabei ZUERST allen
     XHR/fetch-Traffic mitschneiden. Aussteller-Verzeichnisse laden ihre Daten
     fast immer aus einer JSON-API -> wir fischen das JSON-Array mit den Firmen
     direkt heraus (vollstaendig, sauber, inkl. Website/Stand).
     Findet der XHR-Sniffer nichts, fallen wir auf DOM-Extraktion per
     CSS-Selektor zurueck (mit Auto-Scroll + "Mehr laden"/Naechste-Seite).

Nutzung:
  pip install -r requirements.txt
  python -m playwright install chromium        # entfaellt in der CCR-Umgebung
  python scrape.py                 # alle Messen
  python scrape.py --only 09,10    # nur bestimmte (Index aus fairs.json, 1-basiert)
  python scrape.py --headful       # Browser sichtbar (Debug)

Voraussetzung: OFFENER Netzwerk-Egress (siehe README.md). In einer reinen
"Package managers only"-Umgebung schlaegt jeder Abruf mit 403 fehl.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse

BASE = Path(__file__).resolve().parent
OUT = BASE / "out"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")

# Feldnamen, die in JSON-APIs typischerweise den Firmennamen tragen
NAME_KEYS = ["name", "companyName", "company", "exhibitorName", "title",
             "displayName", "firmenname", "aussteller", "orgName"]
WEB_KEYS = ["website", "url", "web", "homepage", "www", "link", "domain"]
CITY_KEYS = ["city", "ort", "town", "location", "stadt"]
COUNTRY_KEYS = ["country", "land", "countryName", "nation"]
STAND_KEYS = ["stand", "booth", "hall", "halle", "standNumber", "boothNumber"]


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def log(msg: str) -> None:
    print(msg, flush=True)


def pick(d: dict, keys: list[str]) -> str:
    for k in d:
        if k.lower() in keys and isinstance(d[k], (str, int, float)) and str(d[k]).strip():
            return str(d[k]).strip()
    return ""


def domain_of(url: str) -> str:
    if not url:
        return ""
    if not url.startswith("http"):
        url = "http://" + url
    try:
        net = urlparse(url).netloc.lower()
        return net[4:] if net.startswith("www.") else net
    except Exception:
        return ""


def looks_like_exhibitor_list(arr) -> bool:
    """Heuristik: Liste von Objekten, von denen >=60% ein Namensfeld haben."""
    if not isinstance(arr, list) or len(arr) < 3:
        return False
    objs = [x for x in arr if isinstance(x, dict)]
    if len(objs) < max(3, len(arr) * 0.6):
        return False
    with_name = sum(1 for o in objs if pick(o, NAME_KEYS))
    return with_name >= len(objs) * 0.6


def find_arrays(node, out: list, depth: int = 0):
    """Rekursiv alle Listen finden, die nach Ausstellern aussehen."""
    if depth > 8:
        return
    if isinstance(node, list):
        if looks_like_exhibitor_list(node):
            out.append(node)
        for x in node:
            find_arrays(x, out, depth + 1)
    elif isinstance(node, dict):
        for v in node.values():
            find_arrays(v, out, depth + 1)


def rows_from_json_array(arr, fair) -> list[dict]:
    rows = []
    for o in arr:
        if not isinstance(o, dict):
            continue
        name = pick(o, NAME_KEYS)
        if not name:
            continue
        web = pick(o, WEB_KEYS)
        rows.append({
            "messe": fair["name"],
            "ort": fair["city"],
            "aussteller": name,
            "website": web,
            "domain": domain_of(web),
            "stadt": pick(o, CITY_KEYS),
            "land": pick(o, COUNTRY_KEYS),
            "halle_stand": pick(o, STAND_KEYS),
            "quelle": fair["url"],
        })
    return rows


def dedupe(rows: list[dict]) -> list[dict]:
    seen, out = set(), []
    for r in rows:
        key = re.sub(r"\s+", " ", r["aussteller"].lower()).strip()
        if key and key not in seen:
            seen.add(key)
            out.append(r)
    return out


# --------------------------------------------------------------------------- #
# method = download  (XLS / CSV / PDF direkt)
# --------------------------------------------------------------------------- #
def scrape_download(fair) -> list[dict]:
    import requests
    url = fair["download_url"]
    log(f"    [download] {url}")
    r = requests.get(url, headers={"User-Agent": UA}, timeout=60)
    r.raise_for_status()
    ct = r.headers.get("content-type", "").lower()
    data = r.content

    # PDF
    if url.lower().endswith(".pdf") or "pdf" in ct:
        return rows_from_pdf(data, fair)
    # Excel
    if url.lower().endswith((".xls", ".xlsx")) or "excel" in ct or "spreadsheet" in ct:
        return rows_from_excel(data, fair)
    # CSV
    if url.lower().endswith(".csv") or "csv" in ct:
        return rows_from_csv(data.decode("utf-8", "replace"), fair)
    # Unbekannt -> Excel versuchen, dann CSV
    try:
        return rows_from_excel(data, fair)
    except Exception:
        return rows_from_csv(data.decode("utf-8", "replace"), fair)


def rows_from_excel(data: bytes, fair) -> list[dict]:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        # aeltere .xls -> xlrd/pandas
        import pandas as pd
        df = pd.read_excel(io.BytesIO(data))
        return rows_from_dataframe(df, fair)
    ws = wb.active
    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    name_i = _col_index(header, NAME_KEYS + ["firma", "unternehmen", "exhibitor"])
    web_i = _col_index(header, WEB_KEYS + ["internet", "webseite"])
    city_i = _col_index(header, CITY_KEYS)
    country_i = _col_index(header, COUNTRY_KEYS)
    stand_i = _col_index(header, STAND_KEYS)
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if name_i is None or name_i >= len(r) or not r[name_i]:
            continue
        web = str(r[web_i]).strip() if web_i is not None and web_i < len(r) and r[web_i] else ""
        rows.append({
            "messe": fair["name"], "ort": fair["city"],
            "aussteller": str(r[name_i]).strip(), "website": web, "domain": domain_of(web),
            "stadt": _cell(r, city_i), "land": _cell(r, country_i),
            "halle_stand": _cell(r, stand_i), "quelle": fair.get("download_url", fair["url"]),
        })
    return rows


def rows_from_dataframe(df, fair) -> list[dict]:
    cols = {c.lower(): c for c in df.columns}
    def col(keys):
        for k in cols:
            if k in keys:
                return cols[k]
        return None
    nc = col(NAME_KEYS + ["firma", "unternehmen", "exhibitor"])
    wc = col(WEB_KEYS + ["internet", "webseite"])
    rows = []
    for _, row in df.iterrows():
        if not nc or not str(row.get(nc, "")).strip() or str(row.get(nc)) == "nan":
            continue
        web = str(row.get(wc, "")).strip() if wc else ""
        web = "" if web == "nan" else web
        rows.append({
            "messe": fair["name"], "ort": fair["city"],
            "aussteller": str(row[nc]).strip(), "website": web, "domain": domain_of(web),
            "stadt": "", "land": "", "halle_stand": "",
            "quelle": fair.get("download_url", fair["url"]),
        })
    return rows


def rows_from_csv(text: str, fair) -> list[dict]:
    dialect = csv.Sniffer().sniff(text[:2048], delimiters=";,\t")
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    rows = []
    for d in reader:
        low = {k.lower().strip(): v for k, v in d.items() if k}
        name = next((low[k] for k in low if k in NAME_KEYS + ["firma", "unternehmen"]), "")
        if not name:
            continue
        web = next((low[k] for k in low if k in WEB_KEYS), "")
        rows.append({
            "messe": fair["name"], "ort": fair["city"],
            "aussteller": name.strip(), "website": web, "domain": domain_of(web),
            "stadt": "", "land": "", "halle_stand": "",
            "quelle": fair.get("download_url", fair["url"]),
        })
    return rows


def rows_from_pdf(data: bytes, fair) -> list[dict]:
    """PDF-Ausstellerverzeichnis: Zeilen mit Firmennamen extrahieren.
    Heuristik + optionale manuelle Nachbearbeitung noetig."""
    try:
        import pdfplumber
    except ImportError:
        log("    [pdf] pdfplumber nicht installiert -> uebersprungen")
        return []
    rows, seen = [], set()
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            for line in (page.extract_text() or "").splitlines():
                line = line.strip()
                # Firmennamen enthalten typ. Rechtsform oder Grossbuchstaben, kein reiner Stand-Code
                if len(line) < 3 or re.fullmatch(r"[\d\.\s\-/A-H]+", line):
                    continue
                m = re.match(r"^(.*?(?:GmbH|AG|KG|e\.?V\.?|Ltd|Inc|B\.?V\.?|GbR|OHG|SE|UG|s\.?r\.?l).*?)(?:\s{2,}|\s+[A-H]\d)", line)
                cand = (m.group(1) if m else line).strip()
                if cand and cand.lower() not in seen and len(cand) < 90:
                    seen.add(cand.lower())
                    rows.append({
                        "messe": fair["name"], "ort": fair["city"], "aussteller": cand,
                        "website": "", "domain": "", "stadt": "", "land": "",
                        "halle_stand": "", "quelle": fair["download_url"],
                    })
    log(f"    [pdf] {len(rows)} Kandidaten (bitte stichprobenartig pruefen)")
    return rows


def _col_index(header, keys):
    for i, h in enumerate(header):
        if h in keys:
            return i
    return None


def _cell(row, i):
    return str(row[i]).strip() if i is not None and i < len(row) and row[i] else ""


# --------------------------------------------------------------------------- #
# method = browser  (Playwright: XHR-Sniffing + DOM-Fallback)
# --------------------------------------------------------------------------- #
def scrape_browser(fair, headful=False) -> list[dict]:
    from playwright.sync_api import sync_playwright

    captured: list = []  # JSON-Bodies aus XHR/fetch

    def on_response(resp):
        try:
            ct = resp.headers.get("content-type", "")
            if "json" not in ct:
                return
            body = resp.json()
            captured.append(body)
        except Exception:
            pass

    rows: list[dict] = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not headful)
        ctx = browser.new_context(user_agent=UA, locale="de-DE")
        page = ctx.new_page()
        page.on("response", on_response)
        log(f"    [browser] lade {fair['url']}")
        page.goto(fair["url"], wait_until="networkidle", timeout=90_000)
        _accept_cookies(page)
        _load_everything(page)  # scroll + "mehr laden" + naechste Seite

        # 1) XHR-JSON auswerten
        arrays: list = []
        for body in captured:
            find_arrays(body, arrays)
        best = max(arrays, key=len) if arrays else None
        if best:
            rows = rows_from_json_array(best, fair)
            log(f"    [browser] XHR-Treffer: {len(rows)} Aussteller aus JSON-API")

        # 2) DOM-Fallback
        if len(rows) < 3:
            rows = _dom_extract(page, fair)
            log(f"    [browser] DOM-Fallback: {len(rows)} Aussteller")

        browser.close()
    return rows


def _accept_cookies(page):
    for sel in [
        "button:has-text('Akzeptieren')", "button:has-text('Alle akzeptieren')",
        "button:has-text('Accept all')", "button:has-text('Accept')",
        "button:has-text('Zustimmen')", "#onetrust-accept-btn-handler",
        "button[aria-label*='akzeptier']",
    ]:
        try:
            el = page.query_selector(sel)
            if el and el.is_visible():
                el.click(timeout=3000)
                page.wait_for_timeout(800)
                return
        except Exception:
            pass


def _load_everything(page, max_rounds=60):
    """Infinite-Scroll + 'Mehr laden'/'Naechste Seite' generisch abarbeiten."""
    last_h = 0
    for _ in range(max_rounds):
        clicked = False
        for sel in [
            "button:has-text('Mehr laden')", "button:has-text('Load more')",
            "button:has-text('mehr anzeigen')", "a:has-text('Weiter')",
            "a[rel='next']", ".pagination-next:not(.disabled) a", "button:has-text('Next')",
        ]:
            try:
                el = page.query_selector(sel)
                if el and el.is_visible():
                    el.click(timeout=3000)
                    page.wait_for_timeout(1500)
                    clicked = True
                    break
            except Exception:
                pass
        page.mouse.wheel(0, 20000)
        page.wait_for_timeout(1000)
        h = page.evaluate("document.body.scrollHeight")
        if h == last_h and not clicked:
            break
        last_h = h


def _dom_extract(page, fair) -> list[dict]:
    names: list[str] = []
    sel = fair.get("selector", "")
    if sel:
        for el in page.query_selector_all(sel):
            t = (el.inner_text() or "").strip()
            if 2 < len(t) < 90:
                names.append(t)
    # Generische Rechtsform-Heuristik, falls Selektor nichts liefert
    if len(names) < 3:
        txt = page.evaluate("document.body.innerText") or ""
        for line in txt.splitlines():
            line = line.strip()
            if re.search(r"\b(GmbH|AG|KG|e\.?V\.?|Ltd|Inc|B\.?V\.?|GbR|SE|UG|s\.?r\.?l)\b", line) and len(line) < 90:
                names.append(line)
    out, seen = [], set()
    for n in names:
        k = n.lower()
        if k not in seen:
            seen.add(k)
            out.append({
                "messe": fair["name"], "ort": fair["city"], "aussteller": n,
                "website": "", "domain": "", "stadt": "", "land": "",
                "halle_stand": "", "quelle": fair["url"],
            })
    return out


# --------------------------------------------------------------------------- #
# Output
# --------------------------------------------------------------------------- #
FIELDS = ["messe", "ort", "aussteller", "website", "domain", "stadt", "land", "halle_stand", "quelle"]


def write_csv(path: Path, rows: list[dict]):
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(path: Path, rows: list[dict]):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aussteller"
    ws.append(FIELDS)
    for r in rows:
        ws.append([r.get(k, "") for k in FIELDS])
    wb.save(path)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", help="Kommagetrennte 1-basierte Indizes, z.B. 9,10")
    ap.add_argument("--headful", action="store_true")
    args = ap.parse_args()

    cfg = json.loads((BASE / "fairs.json").read_text(encoding="utf-8"))
    fairs = cfg["fairs"]
    if args.only:
        want = {int(x) for x in args.only.split(",")}
        fairs = [f for i, f in enumerate(fairs, 1) if i in want]

    OUT.mkdir(exist_ok=True)
    combined: list[dict] = []
    summary = []
    for i, fair in enumerate(fairs, 1):
        log(f"\n[{i}/{len(fairs)}] {fair['name']} ({fair['city']}) -> {fair['method']}")
        rows = []
        try:
            if fair["method"] == "download":
                try:
                    rows = scrape_download(fair)
                except Exception as e:
                    log(f"    download fehlgeschlagen ({e}); Browser-Fallback")
                    rows = scrape_browser(fair, args.headful)
            else:
                rows = scrape_browser(fair, args.headful)
        except Exception as e:
            log(f"    FEHLER: {e}")
        rows = dedupe(rows)
        write_csv(OUT / f"{fair['id']}.csv", rows)
        combined += rows
        summary.append((fair["name"], len(rows)))
        log(f"    -> {len(rows)} Aussteller | out/{fair['id']}.csv")

    combined = dedupe(combined) if False else combined  # Messe-uebergreifend NICHT deduplizieren
    write_csv(OUT / "alle_aussteller.csv", combined)
    try:
        write_xlsx(OUT / "alle_aussteller.xlsx", combined)
    except Exception as e:
        log(f"xlsx uebersprungen: {e}")

    log("\n==================== ZUSAMMENFASSUNG ====================")
    for name, n in summary:
        log(f"  {n:>5}  {name}")
    log(f"  {sum(n for _, n in summary):>5}  GESAMT")
    log(f"\nDateien in: {OUT}")


if __name__ == "__main__":
    main()
